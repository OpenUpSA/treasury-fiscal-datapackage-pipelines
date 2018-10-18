"""
`<province code> - EPRE - <financial year hyphenated> - Final.xlsm` e.g. _"NC - EPRE - 2017-18 - Final.xlsm"_

- Province - code in filename
- Budget year - hyphenated in filename
- List of department names with the sheet label (integers) with its data - sheet "Settings"
- For each department (on their integer-indexed sheet e.g. Education on sheet "1"
  - One-liner description of department C6 which doesn't match what's in the EPRE
  - Table with row for each programme B9 to B28
    - FY-4, FY-3, FY-2 Outcome
    - FY-1 Main Appropriation, Adjusted appropriation, Revised estimate
    - FY
      - Indicative baseline
      - Reprioritised baseline
      - Revised baseline
    - FY+1
      - Indicative baseline
      - Reprioritised baseline
      - Revised baseline
    - FY+2 Indicative baseline
  - Department-level breakdown by economic classification 1, 2, 3, (4 sometimes)
  - For each programme
    - name
    - One-liner description of programme which doesn't match what's in the EPRE (not always available)
    - Table with row for each sub-programme
      - totals by sub-programme
    - Programme-level breakdon by economic classification 1, 2, 3, (4 sometimes
"""

from openpyxl import load_workbook
import csv
import re
import os
import sys
from tempfile import mkdtemp
import shutil
import subprocess

prov_abbrev = {
    'EC': 'Eastern Cape',
    'FS': 'Free State',
    'GT': 'Gauteng',
    'KZN': 'KwaZulu-Natal',
    'LIM': 'Limpopo',
    'MPU': 'Mpumalanga',
    'NC': 'Northern Cape',
    'NW': 'North West',
    'WC': 'Western Cape',
}


def fin_year_str(year):
    return '%d-%d' % (year, (year+1)-2000)


def scrape_file(filename, province_code, budget_financial_year):
    print("%s\n" % filename)
    wb = load_workbook(filename=filename, data_only=True)

    ws = wb['Settings']

    department_sheets = {}

    for row in ws.iter_rows(min_row=13, max_row=13+20):
        if row[4].value:
            sheet_name = row[3].value
            department_name = row[4].value.strip()
            department_sheets[sheet_name] = department_name
            print(sheet_name, department_name)

    csv_filename = 'epre-subprogrammes-%s-%s.csv' % (fin_year_str(budget_financial_year), prov_abbrev[province_code])
    with open(csv_filename, 'w') as csv_file:
        fieldnames = [
            'department',
            'programme_number',
            'programme',
            'subprogramme',
            'financial_year',
            'phase',
            'amount',
        ]
        writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
        writer.writeheader()
        for sheet_name, department_name in department_sheets.items():
            ws = wb[str(sheet_name)]
            rows = {}

            print("\n", sheet_name, department_name)

            # skip to programmes
            for idx, row in enumerate(ws.iter_rows()):
                if row[1].value == 'PROGRAMME DETAILS':
                    programmes_row_idx = idx+2
                    break

            in_subprogs = False
            programme_number = 0
            for idx, row in enumerate(ws.iter_rows(min_row=programmes_row_idx)):
                if ws.row_dimensions[programmes_row_idx + idx].hidden:
                    print("hidden ", end='', flush=True)
                    continue

                ###############################################################
                ## Figure out which section of the PROGRAMME DETAILS section we're in

                if (row[1].value
                    and row[1].font
                    and row[1].font.color
                    and row[1].font.color.rgb == 'FF0000FF'):
                    if row[1].value not in ('Direct Charges'):
                        programme_name = row[1].value.strip()
                        programme_number += 1

                if row[1].value == "R 000's":
                    in_subprogs = True
                    continue
                if row[1].value == 'Total':
                    in_subprogs = False
                    programme_name = None

                ##==============================================================
                ## In the Sub-programme section of a Programme

                if in_subprogs and programme_name:
                    subprogramme_name = row[1].value
                    if not subprogramme_name:
                        continue
                    subprogramme_name = subprogramme_name.strip()

                    rows_key = (
                        department_name,
                        programme_number,
                        programme_name,
                        subprogramme_name,
                    )

                    print("Adding ", rows_key)
                    rows[rows_key] = []

                    columns = {
                        3: (-4, 'Audited Outcome'),
                        4: (-3, 'Audited Outcome'),
                        5: (-2, 'Audited Outcome'),
                        6: (-1, 'Main appropriation'),
                        7: (-1, 'Adjusted appropriation'),
                        8: (-1, 'Revised estimate'),
                        16: (0, 'Main appropriation'),
                        25: (1, 'Medium Term Estimates'),
                        27: (2, 'Medium Term Estimates'),
                    }

                    for col_idx, col_meta in columns.items():
                        year_offset, phase = col_meta
                        amount = row[col_idx-1].value
                        if amount:
                            amount = round(amount) * 1000
                        financial_year = budget_financial_year+year_offset
                        year_phase = {
                            'phase': phase,
                            'financial_year': financial_year,
                            'amount': amount,
                        }
                        rows[rows_key].append(year_phase)

                    # End


            ## Done looping over PROGRAMME DETAILS rows
            ####################################################################

            for row_key, row in rows.items():
                for year_phase in row:
                    year_phase['department'] = row_key[0]
                    year_phase['programme_number'] = row_key[1]
                    year_phase['programme'] = row_key[2]
                    year_phase['subprogramme'] = row_key[3]
                    writer.writerow(year_phase)
            csv_file.flush()


directory = sys.argv[1]

for filename in os.listdir(directory):
    match = re.search('(?P<province_code>[A-Z]{2,3}) - EPRE - (?P<financial_year>\d{4})-\d{2} - Final.xlsm$', filename)
    if match:
        province_code = match.group('province_code')
        budget_financial_year = int(match.group('financial_year'))
        filepath = os.path.join(directory, filename)
        try:
            scrape_file(filepath, province_code, budget_financial_year)
        except ValueError as e:
            print(e)
            print("\nfixing xlsx")
            tempdir = mkdtemp(prefix="epre-scrape")
            tempfilepath = os.path.join(tempdir, filename)
            try:
                print("copying %s to %s" % (filepath, tempfilepath))
                shutil.copyfile(filepath, tempfilepath)

                zip_dir = os.path.join(tempdir, "zip")
                print("unzipping %s to %s" % (tempfilepath, zip_dir))
                exit_code = subprocess.call(["unzip", "-d", zip_dir, tempfilepath])
                if exit_code: raise Exception(exit_code)

                workbook_xml = os.path.join(zip_dir, 'xl/workbook.xml')
                print("removing protection in %s" % workbook_xml)
                exit_code = subprocess.call(["perl", "-pi.bak", "-e", 's/<fileSharing[^>]+>//', workbook_xml])
                if exit_code: raise Exception(exit_code)

                print("zipping %s into %s" % (zip_dir, tempfilepath))
                exit_code = subprocess.call(["zip", "-r", tempfilepath, "'[Content_Types].xml'", "docProps/", "_rels/", "xl/"], cwd=zip_dir)
                if exit_code: raise Exception(exit_code)

                print("Trying scraping again from %s" % tempfilepath)
                scrape_file(tempfilepath, province_code, budget_financial_year)
            finally:
                shutil.rmtree(tempdir)
